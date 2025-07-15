#!/usr/bin/env python3
import json
import requests
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import html
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

RED = '\033[91m'
GREEN = '\033[92m'
YELLOW = '\033[93m'
BEBEBLUE = '\033[94m'
VIOLET = '\033[95m'
CYAN = '\033[96m'
WHITE = '\033[97m'
ENDC = '\033[0m'
BOLD = '\033[1m'

def display_banner():
    print(f"{BEBEBLUE}" + "="*80 + f"{ENDC}")
    print(f"{VIOLET}{BOLD}████████╗██╗  ██╗██████╗ ███████╗ █████╗ ████████╗{ENDC}")
    print(f"{VIOLET}{BOLD}╚══██╔══╝██║  ██║██╔══██╗██╔════╝██╔══██╗╚══██╔══╝{ENDC}")
    print(f"{VIOLET}{BOLD}   ██║   ███████║██████╔╝█████╗  ███████║   ██║   {ENDC}")
    print(f"{VIOLET}{BOLD}   ██║   ██╔══██║██╔══██╗██╔══╝  ██╔══██║   ██║   {ENDC}")
    print(f"{VIOLET}{BOLD}   ██║   ██║  ██║██║  ██║███████╗██║  ██║   ██║   {ENDC}")
    print(f"{VIOLET}{BOLD}   ╚═╝   ╚═╝  ╚═╝╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝   ╚═╝   {ENDC}")
    print()
    print(f"{VIOLET}{BOLD}███╗   ███╗ █████╗ ██████╗ ██████╗ ██╗███╗   ██╗ ██████╗ {ENDC}")
    print(f"{VIOLET}{BOLD}████╗ ████║██╔══██╗██╔══██╗██╔══██╗██║████╗  ██║██╔════╝ {ENDC}")
    print(f"{VIOLET}{BOLD}██╔████╔██║███████║██████╔╝██████╔╝██║██╔██╗ ██║██║  ███╗{ENDC}")
    print(f"{VIOLET}{BOLD}██║╚██╔╝██║██╔══██║██╔═══╝ ██╔═══╝ ██║██║╚██╗██║██║   ██║{ENDC}")
    print(f"{VIOLET}{BOLD}██║ ╚═╝ ██║██║  ██║██║     ██║     ██║██║ ╚████║╚██████╔╝{ENDC}")
    print(f"{VIOLET}{BOLD}╚═╝     ╚═╝╚═╝  ╚═╝╚═╝     ╚═╝     ╚═╝╚═╝  ╚═══╝ ╚═════╝ {ENDC}")
    print()
    print(f"{VIOLET}{BOLD}██████╗ ██████╗  ██████╗ {ENDC}")
    print(f"{VIOLET}{BOLD}██╔══██╗██╔══██╗██╔═══██╗{ENDC}")
    print(f"{VIOLET}{BOLD}██████╔╝██████╔╝██║   ██║{ENDC}")
    print(f"{VIOLET}{BOLD}██╔═══╝ ██╔══██╗██║   ██║{ENDC}")
    print(f"{VIOLET}{BOLD}██║     ██║  ██║╚██████╔╝{ENDC}")
    print(f"{VIOLET}{BOLD}╚═╝     ╚═╝  ╚═╝ ╚═════╝ {ENDC}")
    print()
    print(f"{YELLOW}Advanced MITRE ATT&CK framework analyzer for APT group intelligence{ENDC}")
    print(f"{YELLOW}Maps threat actor techniques, tactics, and procedures with detailed analysis{ENDC}")
    print()
    print(f"{GREEN}{BOLD}Created by Muhap Yahia{ENDC}")
    print(f"{BEBEBLUE}" + "="*80 + f"{ENDC}")

class MITREAnalyzer:
    def __init__(self):
        self.enterprise_url = "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json"
        self.groups = {}
        self.techniques = {}
        self.relationships = []
        self.tactics = {}
        self.country_targets_file = "country_targets.json"
        
    def load_mitre_data(self):
        print(f"{BEBEBLUE}[*] Loading MITRE ATT&CK Enterprise data...{ENDC}")
        try:
            response = requests.get(self.enterprise_url, timeout=30)
            response.raise_for_status()
            data = response.json()
            for obj in data['objects']:
                if obj['type'] == 'intrusion-set':
                    group_data = {
                        'id': obj['id'],
                        'name': obj['name'],
                        'description': obj.get('description', ''),
                        'aliases': obj.get('aliases', []),
                        'created': obj.get('created', ''),
                        'modified': obj.get('modified', ''),
                        'attack_id': None
                    }
                    for ref in obj.get('external_references', []):
                        if ref.get('source_name') == 'mitre-attack':
                            group_data['attack_id'] = ref.get('external_id')
                            break
                    self.groups[obj['id']] = group_data
                elif obj['type'] == 'attack-pattern':
                    technique_data = {
                        'id': obj['id'],
                        'name': obj['name'],
                        'description': obj.get('description', ''),
                        'tactics': [],
                        'platforms': obj.get('x_mitre_platforms', []),
                        'data_sources': obj.get('x_mitre_data_sources', []),
                        'detection': obj.get('x_mitre_detection', ''),
                        'is_subtechnique': obj.get('x_mitre_is_subtechnique', False),
                        'attack_id': None
                    }
                    for phase in obj.get('kill_chain_phases', []):
                        if phase.get('kill_chain_name') == 'mitre-attack':
                            technique_data['tactics'].append(phase['phase_name'])
                    for ref in obj.get('external_references', []):
                        if ref.get('source_name') == 'mitre-attack':
                            technique_data['attack_id'] = ref.get('external_id')
                            break
                    self.techniques[obj['id']] = technique_data
                elif obj['type'] == 'relationship':
                    self.relationships.append({
                        'source_ref': obj['source_ref'],
                        'target_ref': obj['target_ref'],
                        'relationship_type': obj['relationship_type'],
                        'description': obj.get('description', ''),
                        'created': obj.get('created', '')
                    })
                elif obj['type'] == 'x-mitre-tactic':
                    tactic_data = {
                        'id': obj['id'],
                        'name': obj['name'],
                        'description': obj.get('description', ''),
                        'short_name': obj.get('x_mitre_shortname', ''),
                        'attack_id': None
                    }
                    for ref in obj.get('external_references', []):
                        if ref.get('source_name') == 'mitre-attack':
                            tactic_data['attack_id'] = ref.get('external_id')
                            break
                    self.tactics[obj['id']] = tactic_data
            print(f"{GREEN}[+] Loaded {len(self.groups)} groups, {len(self.techniques)} techniques, {len(self.relationships)} relationships{ENDC}")
        except requests.RequestException as e:
            print(f"{RED}[-] Error loading MITRE data: {e}{ENDC}")
            raise
        except json.JSONDecodeError as e:
            print(f"{RED}[-] Error parsing MITRE data: {e}{ENDC}")
            raise

    def map_apt_group(self):
        print(f"\n{YELLOW}=== APT GROUP MAPPING & ANALYSIS ==={ENDC}")
        print(f"{GREEN}Enter APT group name, MITRE ID, or alias to analyze{ENDC}")
        print(f"{BEBEBLUE}Examples: G0006, APT1, Lazarus Group, Comment Crew{ENDC}")
        group_input = input(f"{VIOLET}Enter APT group: {ENDC}").strip()
        if not group_input:
            print(f"{RED}[-] Please enter a valid APT group name or ID{ENDC}")
            return
        group_data = self._find_group_enhanced(group_input)
        if not group_data:
            return
        mapped_group = self._map_group_techniques_enhanced(group_data)
        self._display_enhanced_group_analysis(mapped_group)
        navigator_layer = self._generate_navigator_layer(mapped_group)
        safe_name = group_input.replace(' ', '_').replace('/', '_').lower()
        safe_name = ''.join(c for c in safe_name if c.isalnum() or c in '_-')
        output_file = f"{safe_name}_navigator_layer.json"
        self._save_navigator_layer(navigator_layer, output_file)
        print(f"\n{GREEN}[+] Analysis Complete!{ENDC}")
        print(f"{BEBEBLUE}[+] Navigator layer saved: {output_file}{ENDC}")
        print(f"{YELLOW}[+] Import into MITRE ATT&CK Navigator for visualization{ENDC}")

    def _find_group_enhanced(self, group_input):
        print(f"{BEBEBLUE}[*] Searching for group: {group_input}{ENDC}")
        group_input_lower = group_input.lower().strip()
        for group_id, group_data in self.groups.items():
            if group_data.get('attack_id', '').lower() == group_input_lower:
                print(f"{GREEN}[+] Found by MITRE ID: {group_data['attack_id']}{ENDC}")
                return group_data
        for group_id, group_data in self.groups.items():
            if group_data.get('name', '').lower() == group_input_lower:
                print(f"{GREEN}[+] Found by name: {group_data['name']}{ENDC}")
                return group_data
        for group_id, group_data in self.groups.items():
            for alias in group_data.get('aliases', []):
                if alias.lower() == group_input_lower:
                    print(f"{GREEN}[+] Found by alias: {alias}{ENDC}")
                    return group_data
        print(f"{RED}[-] Group not found in MITRE ATT&CK database{ENDC}")
        suggestions = self._suggest_similar_groups(group_input)
        if suggestions:
            print(f"{YELLOW}[?] Did you mean one of these?{ENDC}")
            for suggestion in suggestions[:5]:
                print(f"    {BEBEBLUE}- {suggestion}{ENDC}")
        return None

    def _map_group_techniques_enhanced(self, group_data):
        print(f"{BEBEBLUE}[*] Mapping techniques for {group_data['name']}{ENDC}")
        group_id = group_data['id']
        technique_count = 0
        enhanced_group = group_data.copy()
        enhanced_group['techniques'] = []
        enhanced_group['tactics'] = set()
        enhanced_group['platforms'] = set()
        enhanced_group['data_sources'] = set()
        for relationship in self.relationships:
            if (relationship['source_ref'] == group_id and
                relationship['relationship_type'] == 'uses' and
                relationship['target_ref'] in self.techniques):
                technique_data = self.techniques[relationship['target_ref']]
                technique_count += 1
                last_seen = self._get_technique_last_seen(relationship['target_ref'])
                technique_entry = {
                    'attack_id': technique_data['attack_id'],
                    'name': technique_data['name'],
                    'description': self._clean_text(technique_data['description']),
                    'tactics': technique_data['tactics'],
                    'platforms': technique_data.get('platforms', []),
                    'data_sources': technique_data.get('data_sources', []),
                    'detection': self._clean_text(technique_data.get('detection', '')),
                    'is_subtechnique': technique_data.get('is_subtechnique', False),
                    'relationship_description': self._clean_text(relationship['description']),
                    'relationship_created': relationship['created'],
                    'last_seen': last_seen
                }
                enhanced_group['techniques'].append(technique_entry)
                enhanced_group['tactics'].update(technique_data['tactics'])
                enhanced_group['platforms'].update(technique_data.get('platforms', []))
                enhanced_group['data_sources'].update(technique_data.get('data_sources', []))
        print(f"{GREEN}[+] Mapped {technique_count} techniques{ENDC}")
        enhanced_group['tactics'] = sorted(list(enhanced_group['tactics']))
        enhanced_group['platforms'] = sorted(list(enhanced_group['platforms']))
        enhanced_group['data_sources'] = sorted(list(enhanced_group['data_sources']))
        return enhanced_group

    def _display_enhanced_group_analysis(self, mapped_group):
        print(f"\n{BEBEBLUE}" + "="*60 + f"{ENDC}")
        print(f"{VIOLET}DETAILED GROUP ANALYSIS{ENDC}")
        print(f"{BEBEBLUE}" + "="*60 + f"{ENDC}")
        print(f"{YELLOW}Group Profile:{ENDC}")
        print(f"  Name: {GREEN}{mapped_group['name']}{ENDC}")
        print(f"  MITRE ID: {GREEN}{mapped_group.get('attack_id', 'Unknown')}{ENDC}")
        print(f"  Aliases: {GREEN}{', '.join(mapped_group.get('aliases', [])) if mapped_group.get('aliases') else 'None'}{ENDC}")
        print(f"  First Seen: {GREEN}{mapped_group.get('created', 'Unknown')[:10] if mapped_group.get('created') else 'Unknown'}{ENDC}")
        print(f"  Last Updated: {GREEN}{mapped_group.get('modified', 'Unknown')[:10] if mapped_group.get('modified') else 'Unknown'}{ENDC}")
        print(f"\n{YELLOW}Attack Statistics:{ENDC}")
        print(f"  Total Techniques: {GREEN}{len(mapped_group['techniques'])}{ENDC}")
        print(f"  Tactics Covered: {GREEN}{len(mapped_group['tactics'])}{ENDC}")
        print(f"  Platforms Targeted: {GREEN}{len(mapped_group['platforms'])}{ENDC}")
        print(f"  Data Sources: {GREEN}{len(mapped_group.get('data_sources', []))}{ENDC}")
        print(f"\n{YELLOW}Tactics Used:{ENDC}")
        for tactic in mapped_group['tactics']:
            tactic_techniques = [t for t in mapped_group['techniques'] if tactic in t.get('tactics', [])]
            print(f"  {GREEN}{tactic.title()}{ENDC}: {len(tactic_techniques)} techniques")
        print(f"\n{YELLOW}Target Platforms:{ENDC}")
        for platform in mapped_group['platforms']:
            print(f"  {GREEN}{platform}{ENDC}")

    def _generate_navigator_layer(self, mapped_group):
        description = mapped_group.get('description', '')
        if len(description) > 200:
            description = description[:200] + "..."
        layer = {
            "name": f"{mapped_group['name']} ({mapped_group.get('attack_id', 'Unknown')}) - Techniques",
            "versions": {
                "attack": "14",
                "navigator": "4.9.1",
                "layer": "4.5"
            },
            "domain": "enterprise-attack",
            "description": f"Techniques used by {mapped_group['name']} based on MITRE ATT&CK data. {description}",
            "filters": {
                "platforms": mapped_group['platforms'] if mapped_group['platforms'] else ["Windows", "Linux", "macOS"]
            },
            "sorting": 0,
            "layout": {
                "layout": "side",
                "aggregateFunction": "average",
                "showID": True,
                "showName": True,
                "showAggregateScores": False,
                "countUnscored": False,
                "expandedSubtechniques": "annotated"
            },
            "hideDisabled": False,
            "techniques": [],
            "gradient": {
                "colors": ["#ff6666", "#ffe766", "#8ec843"],
                "minValue": 0,
                "maxValue": 100
            },
            "legendItems": [{
                "label": f"Used by {mapped_group['name']}",
                "color": "#fd8d3c"
            }],
            "showTacticRowBackground": False,
            "tacticRowBackground": "#dddddd",
            "selectTechniquesAcrossTactics": True,
            "selectSubtechniquesWithParent": False,
            "metadata": [
                {"name": "Group", "value": f"{mapped_group['name']} ({mapped_group.get('attack_id', 'Unknown')})"},
                {"name": "Aliases", "value": ", ".join(mapped_group.get('aliases', [])) if mapped_group.get('aliases') else "None"},
                {"name": "Total Techniques", "value": str(len(mapped_group['techniques']))},
                {"name": "Generated", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                {"name": "Data Source", "value": "MITRE ATT&CK Enterprise"}
            ],
            "links": [{
                "label": "MITRE ATT&CK Group Page",
                "url": f"https://attack.mitre.org/groups/{mapped_group.get('attack_id', '')}/"
            }]
        }
        for technique in mapped_group['techniques']:
            primary_tactic = technique['tactics'][0] if technique['tactics'] else "execution"
            comment = technique.get('relationship_description', '')
            if len(comment) > 200:
                comment = comment[:200] + "..."
            technique_entry = {
                "techniqueID": technique['attack_id'],
                "tactic": primary_tactic,
                "score": 100,
                "color": "#fd8d3c",
                "comment": f"Used by {mapped_group['name']}. {comment}",
                "enabled": True,
                "metadata": [
                    {"name": "Technique", "value": technique['name']},
                    {"name": "Tactics", "value": ", ".join(technique['tactics']) if technique['tactics'] else "Not specified"},
                    {"name": "Platforms", "value": ", ".join(technique.get('platforms', [])) if technique.get('platforms') else "Not specified"},
                    {"name": "Sub-technique", "value": "Yes" if technique.get('is_subtechnique') else "No"},
                    {"name": "Last Seen", "value": technique.get('last_seen', 'Unknown')}
                ],
                "links": [{
                    "label": "MITRE ATT&CK Technique Page",
                    "url": f"https://attack.mitre.org/techniques/{technique['attack_id'].replace('.', '/')}/"
                }]
            }
            if not technique.get('is_subtechnique'):
                technique_entry["showSubtechniques"] = True
            layer["techniques"].append(technique_entry)
        return layer

    def _save_navigator_layer(self, layer_data, filename):
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(layer_data, f, indent=2, ensure_ascii=False, separators=(',', ': '))
            print(f"{GREEN}[+] Navigator layer saved to: {filename}{ENDC}")
        except Exception as e:
            print(f"{RED}[-] Error saving file: {e}{ENDC}")

    def _suggest_similar_groups(self, group_input):
        suggestions = []
        group_input_lower = group_input.lower()
        for group_data in self.groups.values():
            if group_input_lower in group_data.get('name', '').lower():
                suggestions.append(f"{group_data.get('attack_id', 'Unknown')} - {group_data['name']}")
            for alias in group_data.get('aliases', []):
                if group_input_lower in alias.lower():
                    suggestions.append(f"{group_data.get('attack_id', 'Unknown')} - {group_data['name']} (alias: {alias})")
                    break
        return suggestions[:5]

    def _clean_text(self, text):
        if not text:
            return ""
        text = html.unescape(text)
        text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
        text = ' '.join(text.split())
        return text

    def analyze_technique_prevalence(self):
        print(f"\n{YELLOW}=== TECHNIQUE PREVALENCE ANALYSIS ==={ENDC}")
        tactic_input = input(f"{VIOLET}Enter tactic name (e.g., 'Persistence', 'Defense Evasion'): {ENDC}").strip()
        if not tactic_input:
            print(f"{RED}[-] Please enter a valid tactic name{ENDC}")
            return
        tactic_lower = tactic_input.lower().replace(' ', '-')
        technique_usage = {}
        for rel in self.relationships:
            if (rel['relationship_type'] == 'uses' and 
                rel['target_ref'] in self.techniques):
                technique = self.techniques[rel['target_ref']]
                if tactic_lower in [t.lower().replace(' ', '-') for t in technique['tactics']]:
                    tech_id = technique['attack_id']
                    if tech_id not in technique_usage:
                        technique_usage[tech_id] = {
                            'name': technique['name'],
                            'count': 0,
                            'groups': set()
                        }
                    technique_usage[tech_id]['count'] += 1
                    if rel['source_ref'] in self.groups:
                        technique_usage[tech_id]['groups'].add(self.groups[rel['source_ref']]['name'])
        if not technique_usage:
            print(f"{RED}[-] No techniques found for tactic: {tactic_input}{ENDC}")
            return
        print(f"\n{GREEN}[+] Found {len(technique_usage)} techniques for tactic: {tactic_input}{ENDC}")
        print(f"{BEBEBLUE}" + "-" * 60 + f"{ENDC}")
        sorted_techniques = sorted(technique_usage.items(), key=lambda x: x[1]['count'], reverse=True)
        for tech_id, data in sorted_techniques[:10]:
            print(f"{YELLOW}{tech_id} - {data['name']}{ENDC}")
            print(f"  Usage Count: {GREEN}{data['count']}{ENDC}")
            print(f"  Used by: {CYAN}{', '.join(list(data['groups'])[:5])}{ENDC}")
            if len(data['groups']) > 5:
                print(f"  ... and {len(data['groups']) - 5} more groups")
            print()

    def assess_tactic_usage(self):
        print(f"\n{YELLOW}=== TACTIC/TECHNIQUE USAGE ASSESSMENT & EXCEL EXPORT ==={ENDC}")
        technique_input = input(f"{VIOLET}Enter technique name or ID (e.g., 'Registry Run Keys', 'T1547.001'): {ENDC}").strip()
        if not technique_input:
            print(f"{RED}[-] Please enter a valid technique name or ID{ENDC}")
            return
        target_technique = None
        technique_input_lower = technique_input.lower()
        for technique in self.techniques.values():
            if (technique_input_lower == technique['name'].lower() or
                technique_input_lower == technique.get('attack_id', '').lower()):
                target_technique = technique
                break
        if not target_technique:
            print(f"{RED}[-] Technique not found: {technique_input}{ENDC}")
            return
        using_groups = []
        for rel in self.relationships:
            if (rel['relationship_type'] == 'uses' and 
                rel['target_ref'] == target_technique['id'] and
                rel['source_ref'] in self.groups):
                group_data = self.groups[rel['source_ref']]
                using_groups.append({
                    'name': group_data['name'],
                    'attack_id': group_data.get('attack_id', 'Unknown'),
                    'aliases': group_data.get('aliases', []),
                    'description': group_data.get('description', ''),
                    'relationship_description': rel.get('description', ''),
                    'created': group_data.get('created', ''),
                    'modified': group_data.get('modified', ''),
                    'relationship_created': rel.get('created', '')
                })
        choice = input(f"{VIOLET}List all groups or top 20? Enter 'all' or 'top20': {ENDC}").strip().lower()
        if choice == 'top20':
            using_groups.sort(key=lambda x: x.get('relationship_created', ''), reverse=True)
            using_groups = using_groups[:20]
        elif choice != 'all':
            print(f"{YELLOW}[!] Invalid choice, defaulting to all groups.{ENDC}")
        print(f"\n{GREEN}[+] Technique: {target_technique['name']} ({target_technique.get('attack_id', 'Unknown')}){ENDC}")
        print(f"{YELLOW}Tactics: {', '.join(target_technique['tactics'])}{ENDC}")
        print(f"{YELLOW}Platforms: {', '.join(target_technique.get('platforms', []))}{ENDC}")
        print(f"{YELLOW}Used by {len(using_groups)} groups:{ENDC}")
        print(f"{BEBEBLUE}" + "-" * 60 + f"{ENDC}")
        for group in using_groups:
            print(f"{CYAN}{group['attack_id']} - {group['name']}{ENDC}")
            if group['relationship_description']:
                desc = group['relationship_description'][:100] + "..." if len(group['relationship_description']) > 100 else group['relationship_description']
                print(f"  Usage: {desc}")
            print()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Technique Usage Analysis"
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Technique Analysis: {target_technique['name']} ({target_technique.get('attack_id', 'Unknown')})"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws['A2'] = f"Tactics: {', '.join(target_technique['tactics'])}"
            ws['A3'] = f"Platforms: {', '.join(target_technique.get('platforms', []))}"
            ws['A4'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            headers = ["APT Name", "APT Group MITRE ID", "Aliases", "APT Description", "Last Seen", "First Seen", "Usage Description"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            for row_idx, group in enumerate(using_groups, 7):
                first_seen = group.get('created', '')[:10] if group.get('created') else 'Unknown'
                last_seen = group.get('modified', '')[:10] if group.get('modified') else 'Unknown'
                if group.get('relationship_created'):
                    rel_date = group['relationship_created'][:10] if group['relationship_created'] else first_seen
                    first_seen = rel_date
                aliases = ', '.join(group.get('aliases', [])) if group.get('aliases') else 'None'
                description = self._clean_text(group.get('description', ''))
                usage_desc = self._clean_text(group.get('relationship_description', ''))
                ws.cell(row=row_idx, column=1, value=group['name'])
                ws.cell(row=row_idx, column=2, value=group['attack_id'])
                ws.cell(row=row_idx, column=3, value=aliases)
                ws.cell(row=row_idx, column=4, value=description)
                ws.cell(row=row_idx, column=5, value=last_seen)
                ws.cell(row=row_idx, column=6, value=first_seen)
                ws.cell(row=row_idx, column=7, value=usage_desc)
            for col in range(1, 8):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            safe_technique_name = target_technique.get('attack_id', 'unknown').replace('.', '_')
            filename = f"technique_usage_{safe_technique_name}.xlsx"
            wb.save(filename)
            print(f"{GREEN}[+] Excel file created successfully: {filename}{ENDC}")
            print(f"{BEBEBLUE}[+] File contains {len(using_groups)} APT groups using this technique{ENDC}")
        except Exception as e:
            print(f"{RED}[-] Error creating Excel file: {e}{ENDC}")
            print(f"{YELLOW}[!] Make sure you have openpyxl installed: pip install openpyxl{ENDC}")

    def list_top_apt_groups_by_country(self):
        print(f"\n{YELLOW}=== TOP 20 ACTIVE APT GROUPS BY COUNTRY ==={ENDC}")
        print(f"{GREEN}Enter country name to analyze APT groups targeting that region{ENDC}")
        print(f"{BEBEBLUE}Examples: United States, China, Russia, Iran, North Korea{ENDC}")
        country_name = input(f"{VIOLET}Enter country name: {ENDC}").strip()
        if not country_name:
            print(f"{RED}[-] Please enter a valid country name{ENDC}")
            return
        self._save_country_target(country_name)
        print(f"\n{BEBEBLUE}[*] Analyzing APT groups targeting {country_name}...{ENDC}")
        country_lower = country_name.lower()
        current_date = datetime.utcnow()
        one_year_ago = current_date - timedelta(days=365)
        targeting_groups = []
        for group_id, group_data in self.groups.items():
            group_score = 0
            technique_count = 0
            recent_activity = False
            description = group_data.get('description', '').lower()
            if country_lower in description:
                group_score += 10
            for alias in group_data.get('aliases', []):
                if country_lower in alias.lower():
                    group_score += 5
                    break
            for rel in self.relationships:
                if (rel['source_ref'] == group_id and 
                    rel['relationship_type'] == 'uses' and
                    rel['target_ref'] in self.techniques):
                    technique_count += 1
                    try:
                        rel_date = datetime.fromisoformat(rel['created'].replace('Z', '+00:00')).replace(tzinfo=None)
                        if rel_date >= one_year_ago:
                            recent_activity = True
                            group_score += 2
                    except:
                        pass
            if technique_count > 0:
                group_score += min(technique_count, 50)
            if recent_activity:
                group_score += 20
            targeting_score = self._calculate_country_targeting_score(group_data, country_lower)
            group_score += targeting_score
            if group_score > 0:
                targeting_groups.append({
                    'group_data': group_data,
                    'score': group_score,
                    'technique_count': technique_count,
                    'recent_activity': recent_activity,
                    'last_seen': self._get_group_last_activity(group_id)
                })
        targeting_groups.sort(key=lambda x: x['score'], reverse=True)
        top_20_groups = targeting_groups[:20]
        if not top_20_groups:
            print(f"{RED}[-] No APT groups found targeting {country_name}{ENDC}")
            return
        print(f"\n{GREEN}[+] Found {len(targeting_groups)} APT groups with potential targeting of {country_name}{ENDC}")
        print(f"{YELLOW}[+] Displaying top 20 most active groups:{ENDC}")
        print(f"{BEBEBLUE}" + "="*80 + f"{ENDC}")
        for i, group_info in enumerate(top_20_groups, 1):
            group = group_info['group_data']
            score = group_info['score']
            technique_count = group_info['technique_count']
            recent_activity = group_info['recent_activity']
            last_seen = group_info['last_seen']
            print(f"{CYAN}{i:2d}. {group.get('attack_id', 'Unknown'):8s} - {group['name']}{ENDC}")
            print(f"    {YELLOW}Score: {score:3d} | Techniques: {technique_count:3d} | Recent Activity: {'Yes' if recent_activity else 'No'}{ENDC}")
            print(f"    {YELLOW}Last Seen: {last_seen}{ENDC}")
            if group.get('aliases'):
                aliases = ', '.join(group['aliases'][:3])
                if len(group['aliases']) > 3:
                    aliases += f" (+{len(group['aliases']) - 3} more)"
                print(f"    {BEBEBLUE}Aliases: {aliases}{ENDC}")
            description = group.get('description', '')
            if country_lower in description.lower():
                words = description.split()
                relevant_snippet = []
                for j, word in enumerate(words):
                    if country_lower in word.lower():
                        start = max(0, j-5)
                        end = min(len(words), j+6)
                        relevant_snippet = words[start:end]
                        break
                if relevant_snippet:
                    snippet_text = ' '.join(relevant_snippet)
                    if len(snippet_text) > 100:
                        snippet_text = snippet_text[:100] + "..."
                    print(f"    {GREEN}Targeting: {snippet_text}{ENDC}")
            print()
        self._save_country_analysis_results(country_name, top_20_groups)
        print(f"{GREEN}[+] Analysis complete! Results saved to {country_name.lower().replace(' ', '_')}_apt_analysis.json{ENDC}")

    def _save_country_target(self, country_name):
        try:
            country_data = {}
            if os.path.exists(self.country_targets_file):
                with open(self.country_targets_file, 'r', encoding='utf-8') as f:
                    country_data = json.load(f)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if 'queries' not in country_data:
                country_data['queries'] = []
            country_data['queries'].append({
                'country': country_name,
                'timestamp': timestamp
            })
            with open(self.country_targets_file, 'w', encoding='utf-8') as f:
                json.dump(country_data, f, indent=2, ensure_ascii=False)
            print(f"{GREEN}[+] Country target '{country_name}' saved to {self.country_targets_file}{ENDC}")
        except Exception as e:
            print(f"{YELLOW}[!] Warning: Could not save country target: {e}{ENDC}")

    def _calculate_country_targeting_score(self, group_data, country_lower):
        score = 0
        country_keywords = {
            'united states': ['us', 'usa', 'american', 'washington'],
            'china': ['chinese', 'beijing', 'prc'],
            'russia': ['russian', 'moscow', 'kremlin'],
            'iran': ['iranian', 'tehran', 'persian'],
            'north korea': ['dprk', 'pyongyang', 'korean'],
            'south korea': ['rok', 'seoul', 'korean'],
            'israel': ['israeli', 'tel aviv', 'jerusalem'],
            'india': ['indian', 'delhi', 'mumbai'],
            'japan': ['japanese', 'tokyo'],
            'germany': ['german', 'berlin'],
            'france': ['french', 'paris'],
            'united kingdom': ['uk', 'british', 'london', 'england'],
            'ukraine': ['ukrainian', 'kiev', 'kyiv'],
            'taiwan': ['taiwanese', 'taipei']
        }
        description = group_data.get('description', '').lower()
        if country_lower in description:
            score += 15
        if country_lower in country_keywords:
            for keyword in country_keywords[country_lower]:
                if keyword in description:
                    score += 8
                    break
        gov_keywords = ['government', 'military', 'defense', 'ministry', 'embassy', 'diplomatic']
        for keyword in gov_keywords:
            if keyword in description:
                score += 5
                break
        return score

    def _get_group_last_activity(self, group_id):
        latest_date = None
        for rel in self.relationships:
            if rel['source_ref'] == group_id:
                try:
                    rel_date = datetime.fromisoformat(rel['created'].replace('Z', '+00:00')).replace(tzinfo=None)
                    if latest_date is None or rel_date > latest_date:
                        latest_date = rel_date
                except:
                    pass
        if latest_date:
            return latest_date.strftime('%Y-%m-%d')
        else:
            group_data = self.groups.get(group_id, {})
            modified = group_data.get('modified', '')
            if modified:
                try:
                    mod_date = datetime.fromisoformat(modified.replace('Z', '+00:00')).replace(tzinfo=None)
                    return mod_date.strftime('%Y-%m-%d')
                except:
                    pass
        return 'Unknown'

    def _save_country_analysis_results(self, country_name, top_groups):
        try:
            filename = f"{country_name.lower().replace(' ', '_')}_apt_analysis.json"
            results = {
                'country': country_name,
                'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_groups_analyzed': len(self.groups),
                'top_20_groups': []
            }
            for i, group_info in enumerate(top_groups, 1):
                group = group_info['group_data']
                results['top_20_groups'].append({
                    'rank': i,
                    'mitre_id': group.get('attack_id', 'Unknown'),
                    'name': group['name'],
                    'aliases': group.get('aliases', []),
                    'score': group_info['score'],
                    'technique_count': group_info['technique_count'],
                    'recent_activity': group_info['recent_activity'],
                    'last_seen': group_info['last_seen'],
                    'description': group.get('description', '')[:500] + "..." if len(group.get('description', '')) > 500 else group.get('description', '')
                })
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"{YELLOW}[!] Warning: Could not save analysis results: {e}{ENDC}")

    def _get_technique_last_seen(self, technique_id):
        latest_date = None
        for rel in self.relationships:
            if rel['target_ref'] == technique_id:
                try:
                    rel_date = datetime.fromisoformat(rel['created'].replace('Z', '+00:00')).replace(tzinfo=None)
                    if latest_date is None or rel_date > latest_date:
                        latest_date = rel_date
                except:
                    pass
        return latest_date.strftime('%Y-%m-%d') if latest_date else 'Unknown'
    
    def _group_used_tactic_recently(self, group_id, tactic_short_name, cutoff_date, current_date):
        for rel in self.relationships:
            if (rel['source_ref'] == group_id and 
                rel['target_ref'] in self.techniques):
                tech = self.techniques[rel['target_ref']]
                if tactic_short_name in [t.lower().replace(' ', '-') for t in tech['tactics']]:
                    try:
                        rel_date = datetime.fromisoformat(rel['created'].replace('Z', '+00:00')).replace(tzinfo=None)
                        if cutoff_date <= rel_date <= current_date:
                            return True
                    except:
                        pass
        return False
    
    def _group_used_technique_recently(self, group_id, technique, cutoff_date, current_date):
        for rel in self.relationships:
            if (rel['source_ref'] == group_id and 
                rel['target_ref'] in self.techniques and
                self.techniques[rel['target_ref']]['id'] == technique['id']):
                try:
                    rel_date = datetime.fromisoformat(rel['created'].replace('Z', '+00:00')).replace(tzinfo=None)
                    if cutoff_date <= rel_date <= current_date:
                        return True
                except:
                    pass
        return False
    
    def run(self):
        self.load_mitre_data()
        print(f"\n{BEBEBLUE}" + "="*60 + f"{ENDC}")
        print(f"{VIOLET}THREAT MAPPING PRO - MITRE ATT&CK ANALYZER{ENDC}")
        print(f"{GREEN}Advanced Threat Intelligence Analysis{ENDC}")
        print(f"{BEBEBLUE}" + "="*60 + f"{ENDC}")
        while True:
            print(f"\n{YELLOW}ANALYSIS MODES:{ENDC}")
            print(f"{GREEN}1. Map APT Group (by ID, name, or alias){ENDC}")
            print(f"{GREEN}2. Analyze Tactic Prevalence (e.g., 'Persistence'){ENDC}")
            print(f"{GREEN}3. Assess Technique Usage & Export to Excel {ENDC}")
            print(f"{GREEN}4. List Top 20 APT Groups by Country Target{ENDC}")
            print(f"{BEBEBLUE}" + "-" * 60 + f"{ENDC}")
            try:
                choice = int(input(f"{VIOLET}Select option (1-4): {ENDC}"))
                if choice == 1:
                    self.map_apt_group()
                elif choice == 2:
                    self.analyze_technique_prevalence()
                elif choice == 3:
                    self.assess_tactic_usage()
                elif choice == 4:
                    self.list_top_apt_groups_by_country()
                else:
                    print(f"{RED}[-] Please enter 1, 2, 3, or 4{ENDC}")
                    continue
                if input(f"\n{BEBEBLUE}[?] Continue analysis? (y/n): {ENDC}").lower() != 'y':
                    break
            except KeyboardInterrupt:
                print(f"\n{RED}[!] Exiting...{ENDC}")
                break
            except ValueError:
                print(f"{RED}[-] Please enter a valid number{ENDC}")
                continue
            except Exception as e:
                print(f"{RED}[-] Error: {e}{ENDC}")
                continue

if __name__ == "__main__":
    display_banner()
    analyzer = MITREAnalyzer()
    analyzer.run()
